import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")

#print(wb.sheetnames) VER cantidad de hojas

#print(wb["Hoja1"]) acceder

hoja = wb.active
#print(hoja)

wb.create_sheet("Hoja2")
# print(wb.sheetnames)

hoja2 = wb["Hoja2"]
hoja.title = "Nuevo Titulo3"

# print(
#     hoja.max_row,
#     hoja.max_column,
# )
celda =hoja["A2"]
celda.value = "Id Usuario"
# print(celda.value)

celda2 = hoja.cell(row = 6, column=2)
# print(
#     celda2.value,
#     celda2.row,
#     celda2.column,
#     celda2.coordinate,
      
#       )

for fila in range(1,hoja.max_row + 1):
    for columna in range(1,hoja.max_column + 1):
        celda = hoja.cell(row=fila,column=columna)
        #print(fila,columna,celda.value)
        
columna = hoja["B"]
fila = hoja["3"]
# print(columna)
# print(fila)

hoja.append([5,6,7])
# print(hoja.rows)

hoja.delete_rows(1,2)

#Gaurdar Todo

wb.save("Nuevo_exel.xlsx")